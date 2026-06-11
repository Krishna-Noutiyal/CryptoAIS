import openpyxl.cell
import pandas as pd
from dataclasses import dataclass
from math import isnan
import random as r
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import datetime as dt


@dataclass
class ExcelProcessor:

    def _select_workbook(self, file_name: str, sheet_name: str = "Crypto") -> None:
        """
        Opens an existing Excel workbook for Form-16 using openpyxl and selects the specified worksheet.

        This method loads the Excel file provided by `file_name` and sets the `self.workbook` attribute
        to the loaded workbook object. It then selects the worksheet specified by `sheet_name` (defaulting
        to "FORM-16") and assigns it to the `self.ws` attribute for further processing.

        Args:
            file_name (str): The path to the Excel file to open. This should be a valid .xlsx file.
            sheet_name (str, optional): The name of the worksheet to select from the workbook.
            Defaults to "FORM-16".

        Raises:
            FileNotFoundError: If the specified Excel file does not exist.
            KeyError: If the specified worksheet name does not exist in the workbook.

        Side Effects:
            Sets self.workbook to the loaded openpyxl workbook.
            Sets self.ws to the selected worksheet within the workbook.
        """
        self.workbook = openpyxl.load_workbook(file_name)
        self._select_worksheet(self.workbook, sheet_name)

    def _select_worksheet(self, workbook: openpyxl.Workbook, sheet_name: str) -> None:
        """
        Selects a worksheet from an openpyxl workbook by name and assigns it to self.ws.

        This method takes an openpyxl Workbook object and a worksheet name, and sets the
        self.ws attribute to the corresponding worksheet. It is used internally after loading
        a workbook to prepare for further processing or data manipulation.

        Args:
            workbook (openpyxl.Workbook): The loaded Excel workbook object.
            sheet_name (str): The name of the worksheet to select.

        Raises:
            KeyError: If the specified worksheet name does not exist in the workbook.

        Side Effects:
            Sets self.ws to the selected worksheet within the workbook.
        """
        try:
            self.ws = workbook[sheet_name]
            del workbook[sheet_name]
            self.ws = self.workbook.create_sheet(sheet_name, 0)
            # raise FileExistsError("The
        except:
            # Sheet doesn't exist
            self.ws = self.workbook.create_sheet(sheet_name, 0)

    def _apply_style(self, cell, style_key: str) -> None:
        """
        Applies a predefined style from self.s to a given cell.

        Args:
            cell (openpyxl.cell.cell.Cell): The cell to apply the style to.
            style_key (str): The key of the style in self.s to apply.

        Raises:
            KeyError: If the style_key does not exist in self.s.
        """
        if style_key not in self.s:
            raise KeyError(f"Style '{style_key}' not found in self.s")

        style = self.s[style_key]
        for attribute, value in style.items():
            setattr(cell, attribute, value)

    def _add_formats(self) -> None:
        """
        Defines commonly used cell styles for formatting Excel sheets and stores them in self.s.

        Side Effects:
            Sets self.s to a dictionary containing predefined cell styles.
        """
        # Define font styles

        bold_font = Font(name="calibri", bold=True, size=16)
        medium_font = Font(name="calibri", size=18, bold=True, color="FFFFFF")
        medium_bold_font = Font(name="calibri", size=18, bold=True, color="FFFFFF")
        normal_font = Font(name="calibri", size=12)
        big_font = Font(name="calibri", bold=True, size=26)
        black_font = Font(name="calibri", color="FFFFFF", bold=True, size=16)
        black_font_h = Font(name="calibri", color="FFFFFF", bold=True, size=26)

        # Define alignment styles
        center_alignment = Alignment(
            horizontal="center", vertical="center", wrapText=True
        )

        # Define border styles
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Define fill styles for background colors
        dark_red_fill = PatternFill(
            start_color="FF0066", end_color="FF0066", fill_type="solid"
        )
        medium_red_fill = PatternFill(
            start_color="FF3399", end_color="FF3399", fill_type="solid"
        )
        light_red_fill = PatternFill(
            start_color="FF6699", end_color="FF6699", fill_type="solid"
        )
        black_fill = PatternFill(
            start_color="262626", end_color="262626", fill_type="solid"
        )
        blue_fill = PatternFill(
            start_color="002060", end_color="002060", fill_type="solid"
        )
        green_fill = PatternFill(
            start_color="00B050", end_color="00B050", fill_type="solid"
        )
        red_fill = PatternFill(
            start_color="FF5050", end_color="FF5050", fill_type="solid"
        )

        # Store styles in self.s
        self.s = {
            "blank": {"font": normal_font, "alignment": center_alignment},
            "blank_bold": {"font": bold_font, "alignment": center_alignment},
            "dark_red": {
                "font": bold_font,
                "alignment": center_alignment,
                "fill": dark_red_fill,
                "border": thin_border,
            },
            "medium_red": {
                "font": bold_font,
                "alignment": center_alignment,
                "fill": medium_red_fill,
                "border": thin_border,
            },
            "light_red": {
                "font": bold_font,
                "alignment": center_alignment,
                "fill": light_red_fill,
                "border": thin_border,
            },
            "black": {
                "font": black_font,
                "alignment": center_alignment,
                "fill": black_fill,
                "border": thin_border,
            },
            "blue": {
                "font": medium_font,
                "alignment": center_alignment,
                "fill": blue_fill,
                "border": thin_border,
            },
            "green": {
                "font": medium_bold_font,
                "alignment": center_alignment,
                "fill": green_fill,
                "border": thin_border,
            },
            "red": {
                "font": medium_bold_font,
                "alignment": center_alignment,
                "fill": red_fill,
                "border": thin_border,
            },
            "green_h": {
                "font": big_font,
                "alignment": center_alignment,
                "fill": green_fill,
                "border": thin_border,
            },
            "dark_red_h": {
                "font": big_font,
                "alignment": center_alignment,
                "fill": dark_red_fill,
                "border": thin_border,
            },
            "medium_red_h": {
                "font": bold_font,
                "alignment": center_alignment,
                "fill": medium_red_fill,
                "border": thin_border,
            },
            "light_red_h": {
                "font": big_font,
                "alignment": center_alignment,
                "fill": light_red_fill,
                "border": thin_border,
            },
            "black_h": {
                "font": black_font_h,
                "alignment": center_alignment,
                "fill": black_fill,
                "border": thin_border,
            },
        }

    def _set_worksheet_dimensions(self):
        """
        Set all rows to height 43 and all columns to width 26.1
        """
        from openpyxl.utils import get_column_letter

        # Set row heights for all existing rows plus extra buffer
        max_row = max(self.ws.max_row, 1000)  # At least 100 rows
        for row in range(1, max_row + 1):
            self.ws.row_dimensions[row].height = 43

        # Set column widths for all existing columns
        self.ws.sheet_format.defaultColWidth = 30

    def _set_default_style(self, max_rows: int = 200, max_cols: int = 50) -> None:
        """
        Apply default 'blank' style to all cells in the worksheet
        Args:
            max_rows: Number of rows to style (default 200)
            max_cols: Number of columns to style (default 50)
        """
        for row in range(1, max_rows + 1):
            for col in range(1, max_cols + 1):
                cell = self.ws.cell(row=row, column=col)
                self._apply_style(cell, "blank")

    def set(
        self,
        cell: openpyxl.cell.Cell,
        value: str | int | float | dt.datetime,
        style: str,
        type: str = "general",
    ) -> None:
        """
        Sets the cell with value, style, and optionally cell type.

        Args:
            cell (openpyxl.cell.Cell): The cell to set.
            value (str | int | float): The value to assign.
            style (str): The style key to apply.
            type (str, optional): The type of the cell ('general', 'date', 'text', etc.). Defaults to 'general'.
        """
        cell.value = value
        self._apply_style(cell, style)

        # Set cell number format based on type
        if type == "date":
            cell.number_format = "dd/mm/yyyy"
        elif type == "text":
            cell.number_format = "@"
        elif type == "general":
            cell.number_format = "General"
        # You can add more types/formats as needed

    def make_dashboard(
        self, form_16: str, df: pd.DataFrame, sheet_name: str = "Crypto"
    ) -> bool:
        """
        Creates the Dashboard for crypto Calculations
        """

        self._select_workbook(form_16, sheet_name)

        self._add_formats()

        self._set_worksheet_dimensions()
        self._set_default_style()

        self.ws.merge_cells("A1:G1")
        crypto_details = self.ws["A1"]
        self.set(crypto_details, "Crypto Details", "black_h")

        self.ws.column_dimensions["A"].width = 10
        source = self.ws["A2"]
        self.set(source, "S.No", "dark_red")

        doa = self.ws["B2"]
        self.set(doa, "Date of Acquisition", "medium_red")

        dot = self.ws["C2"]
        self.set(dot, "Date of Transfer", "light_red")

        coa = self.ws["D2"]
        self.set(coa, "Cost of Acquisition", "dark_red")

        cr = self.ws["E2"]
        self.set(cr, "Consideration Recived", "medium_red")

        cr = self.ws["F2"]
        self.set(cr, "Capital Gain", "dark_red")

        cg = self.ws["G2"]
        self.set(cg, "TDS Deducted", "blue")

        """################# Evaluationg Crypto Data #################"""

        # Rename columns for consistency
        df = df.rename(
            columns={
                "Information Source": "source",
                "Date of Payment/Credit": "date_of_transfer",
                "Amount Paid/Credited - Reported by Source": "consideration_received",
            }
        )

        """################# Inserting Crypto Data #################"""

        # Inserting Details

        df["date_of_acquisition"] = None
        df["cost_of_acquisition"] = None
        df["capital_gain"] = None
        df["tds_deducted"] = None

        df_values = df.values
        for i, data in enumerate(df_values):

            # Cell Number
            cno = str(3 + i)

            print(f"({i}) \t {data[0]} \t {data[1].date()} \t {data[2]}")

            # Inserting Source Serial No ( sno )
            sno_cell_index = "A" + cno
            sno_cell = self.ws[sno_cell_index]
            self.set(sno_cell, i + 1, "blank")

            # Inserting Date of Transfer ( dot )
            dot_cell_index = "C" + cno
            dot_cell = self.ws[dot_cell_index]
            self.set(dot_cell, data[1], "blank", "date")

            # Inserting Consideration Received
            cr_cell_index = "E" + cno
            cr_cell = self.ws[cr_cell_index]
            cr_value = data[2]
            self.set(cr_cell, cr_value, "blank")

            """################# Generating Crypto Details #################"""

            # Generating Date of Acquisition ( doa )
            # Generate a random date between 01-04-2024 and data[1]
            start_date = dt.datetime(2024, 4, 1)
            end_date = data[1]
            if end_date <= start_date:
                random_date = start_date
            else:
                delta = end_date - start_date
                random_days = r.randint(0, delta.days)
                random_date = start_date + dt.timedelta(days=random_days)

            doa_cell = "B" + cno
            doa_cell_obj = self.ws[doa_cell]
            self.set(doa_cell_obj, random_date, "blank", "date")

            # Inserting doa in the original dataframe
            df.at[i, "date_of_acquisition"] = random_date

            # Generate Cost of Acquisition (coa) - mostly greater than consideration_received (data[2])
            # 80% chance to be greater, 20% chance to be less or equal
            if r.random() < 0.8:
                # Greater: add 5% to 30% random premium
                increment = r.uniform(0.05, 0.3)
                coa_value = round(data[2] * (1 + increment), 2)
            else:
                # Less or equal: subtract up to 20%
                decrement = r.uniform(0, 0.2)
                coa_value = round(data[2] * (1 - decrement), 2)

            coa_cell_index = "D" + cno
            coa_cell = self.ws[coa_cell_index]
            self.set(coa_cell, int(coa_value), "blank")

            # Inserting coa in the original dataframe
            df.at[i, "cost_of_acquisition"] = coa_value

            # Calculate Capital Gain (cg): consideration_received - cost_of_acquisition
            cg_value = cr_value - coa_value
            df.at[i, "capital_gain"] = cg_value

            cg_cell_index = "F" + cno
            cg_cell = self.ws[cg_cell_index]

            cg_cell_formula = "=E" + cno + "-D" + cno + ""
            self.set(cg_cell, cg_cell_formula, "blank")

            # Inserting TDS Deducted in the original dataframe
            df.at[i, "tds_deducted"] = max(cg_value, 0)

            tds_cell_index = "G" + cno
            tds_cell = self.ws[tds_cell_index]

            tds_cell_formula = f"=MAX(F{cno},0)"
            self.set(tds_cell, tds_cell_formula, "blank")

        """################# Total Crypto Gain/Loss #################"""
        # Calculate the number of data rows
        num_rows = len(df)
        last = num_rows + 3

        cr = sum(df["consideration_received"])
        cg = sum(df["capital_gain"])

        # Total Heading
        self.ws.merge_cells(f"A{last}:C{last}")
        self.set(self.ws[f"A{last}"], "Total", "black_h")

        # Coa total ( Cost of Acquisition )
        coa_total = f"=SUM(D3:D{last-1})"
        self.set(self.ws[f"D{last}"], coa_total, "red")

        # Cr total ( Consideration Recived )
        cr_total = f"=SUM(E3:E{last-1})"
        self.set(self.ws[f"E{last}"], cr_total, "red")

        # cg total ( Consideration Recived )
        cg_total = f"=SUM(F3:F{last-1})"
        self.set(self.ws[f"F{last}"], cg_total, "red")

        # tds total ( Consideration Recived )
        tds_total = f"=SUM(G3:G{last-1})"
        self.set(self.ws[f"G{last}"], tds_total, "red")

        for ws in self.workbook.worksheets:
            ws.sheet_view.tabSelected = None

        self.workbook.active = self.ws
        self.workbook.close()
        self.workbook.save(form_16)

        return True


if __name__ == "__main__":
    # Create an instance of CSVProcessor
    test = ExcelProcessor()

    from csv_processor import CSVProcessor
    import glob
    import os

    test_folder = "test/cryptodata"
    file_list = glob.glob(os.path.join(test_folder, "*.csv"))
    print("Files found:", file_list)
    df = CSVProcessor().combine_csvs(file_list)

    test.make_dashboard("test/cryptodata/Form-16 .xlsx", df)
